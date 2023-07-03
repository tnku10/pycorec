import csv  # system included
import ctypes  # system included
import datetime  # system included
from itertools import zip_longest  # system included
from pathlib import Path  # system included
import re  # system included
from typing import Callable  # system included
from typing import Union  # system included
import customtkinter  # pip install customtkinter
from natsort import natsorted  # pip install natsort
from openpyxl.styles import Font  # pip install openpyxl
from openpyxl import Workbook  # pip install openpyxl
from PIL import Image, ImageTk  # pip install pillow


pycorec_version = "2.0.2"


def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])


class FloatSpinbox(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: Union[int, float] = 1,
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)

        self.step_size = step_size
        self.command = command

        self.configure(fg_color=("gray78", "gray28"))  # set frame color

        self.grid_columnconfigure(2, weight=0)  # buttons don't expand
        self.grid_columnconfigure(1, weight=1)  # entry expands

        self.subtract_button = customtkinter.CTkButton(self, text="-", width=height-6, height=height-6,
                                                       command=self.subtract_button_callback)
        self.subtract_button.grid(row=0, column=0, padx=(3, 0), pady=3)

        self.entry = customtkinter.CTkEntry(self, width=width-(2*height), height=height-6, border_width=0)
        self.entry.grid(row=0, column=1, columnspan=1, padx=3, pady=3, sticky="ew")

        self.add_button = customtkinter.CTkButton(self, text="+", width=height-6, height=height-6,
                                                  command=self.add_button_callback)
        self.add_button.grid(row=0, column=2, padx=(0, 3), pady=3)

        # default value
        self.entry.insert(0, "")

    def add_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = float(self.entry.get()) + self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def subtract_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = float(self.entry.get()) - self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def get(self) -> Union[float, None]:
        try:
            return float(self.entry.get())
        except ValueError:
            return None

    def set(self, value: float):
        self.entry.delete(0, "end")
        self.entry.insert(0, str(float(value)))


class ArrowButton(customtkinter.CTkFrame):
    def __init__(self, *args,
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, **kwargs)

        self.command = command

        self.configure(fg_color="#333333")

        self.up_button = customtkinter.CTkButton(self, text="↑", height=20, width=20,
                                                 command=self.up_button_callback)
        self.up_button.grid(row=0, column=1, padx=1, pady=1)

        self.down_button = customtkinter.CTkButton(self, text="↓", height=20, width=20,
                                                   command=self.down_button_callback)
        self.down_button.grid(row=2, column=1, padx=1, pady=1)

        self.left_button = customtkinter.CTkButton(self, text="←", height=20, width=20,
                                                   command=self.left_button_callback)
        self.left_button.grid(row=1, column=0, padx=1, pady=1)

        self.down_button = customtkinter.CTkButton(self, text="→", height=20, width=20,
                                                   command=self.right_button_callback)
        self.down_button.grid(row=1, column=2, padx=1, pady=1)

    def up_button_callback(self):
        if self.command is not None:
            self.command(dx=0, dy=-10)
            return

    def down_button_callback(self):
        if self.command is not None:
            self.command(dx=0, dy=10)
            return

    def left_button_callback(self):
        if self.command is not None:
            self.command(dx=-10, dy=0)
            return

    def right_button_callback(self):
        if self.command is not None:
            self.command(dx=10, dy=0)
            return


class PyCorec:
    def __init__(self):
        super().__init__()
        self.first_run = True
        self.magnification = 1
        self.photo_image = None
        self.resized_image = None
        self.image_file_name = None
        self.image_height = 1
        self.image_width = 1
        self.resized_image_height = 1
        self.resized_image_width = 1
        self.cm_per_px = None
        self.fps = None
        self.interval = 1
        self.image_paths = []
        self.current_image_index = 0
        self.coordinates = []
        self.pos = []
        self.file_list = []
        self.zoom_level = 1.0
        self.offset_x = 0
        self.offset_y = 0

        customtkinter.set_appearance_mode("Dark")

        self.root = customtkinter.CTk()
        self.root.title(f"PyCorec {pycorec_version}")

        # configure window size
        user32 = ctypes.windll.user32
        screen_width = user32.GetSystemMetrics(0)
        screen_height = user32.GetSystemMetrics(1) - user32.GetSystemMetrics(3)
        self.root.geometry(f"{screen_width}x{screen_height}+0+0")
        self.root.attributes('-topmost', True)

        self.frame = customtkinter.CTkFrame(self.root)
        self.frame.pack(fill=customtkinter.BOTH, expand=True)

        self.image_frame = customtkinter.CTkFrame(self.frame)
        self.image_frame.pack(side=customtkinter.LEFT, fill=customtkinter.BOTH, expand=True)

        self.canvas = customtkinter.CTkCanvas(self.image_frame, bg="white")
        self.canvas.pack(fill=customtkinter.BOTH, expand=True)

        self.bottom_frame = customtkinter.CTkFrame(self.root)
        self.bottom_frame.pack(side=customtkinter.BOTTOM, fill=customtkinter.X)

        self.label_frame = customtkinter.CTkFrame(self.bottom_frame)
        self.label_frame.pack(side=customtkinter.LEFT, fill=customtkinter.X, padx=10, pady=10)

        # labels
        self.coordinates_label = customtkinter.CTkLabel(self.label_frame, text="Coordinates (px): (0, 0)")
        self.coordinates_label.pack(side=customtkinter.LEFT)

        self.frame_interval_label = customtkinter.CTkLabel(self.label_frame, text="Frame Interval: 1")
        self.frame_interval_label.pack(side=customtkinter.LEFT, padx=10)

        self.image_size_label = customtkinter.CTkLabel(self.label_frame, text="Image Size: ")
        self.image_size_label.pack(side=customtkinter.LEFT, padx=10)

        self.resized_image_size_label = customtkinter.CTkLabel(self.label_frame, text="Displayed Image Size: ")
        self.resized_image_size_label.pack(side=customtkinter.LEFT, padx=10)

        self.image_magnification_label = customtkinter.CTkLabel(self.label_frame, text="Image Magnification (%): ")
        self.image_magnification_label.pack(side=customtkinter.LEFT, padx=10)

        self.fps_label = customtkinter.CTkLabel(self.label_frame, text="fps: ")
        self.fps_label.pack(side=customtkinter.LEFT, padx=10)

        self.cm_per_px_label = customtkinter.CTkLabel(self.label_frame, text="cm/px: ")
        self.cm_per_px_label.pack(side=customtkinter.LEFT, padx=10)

        #  buttons
        self.button_frame = customtkinter.CTkFrame(self.frame)
        self.button_frame.pack(side=customtkinter.RIGHT, fill=customtkinter.Y, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Open Images by",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_dir_button = customtkinter.CTkButton(self.button_frame, text="Directory Selection",
                                                             command=self.get_dir)
        self.open_image_dir_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_range_button = customtkinter.CTkButton(self.button_frame, text="Bounded Selection",
                                                               command=self.get_range)
        self.open_image_range_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_click_button = customtkinter.CTkButton(self.button_frame, text="Click Selection",
                                                               command=self.get_files)
        self.open_image_click_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent",
                                                    hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Move Image", fg_color="transparent",
                                                    hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=0)

        self.move_button = ArrowButton(self.button_frame, command=self.move_image)
        self.move_button.pack(fill=customtkinter.X, padx=70, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Image Magnification (%)",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.zoom_spinbox = FloatSpinbox(self.button_frame, width=150, step_size=3)
        self.zoom_spinbox.pack(fill=customtkinter.X, padx=10, pady=0)

        self.zoom_in_button = customtkinter.CTkButton(self.button_frame, text="Apply", command=self.zoom_image)
        self.zoom_in_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.fit_image_to_window_button = customtkinter.CTkButton(self.button_frame, text="Reset to Window Size",
                                                                  command=self.fit_image_to_window)
        self.fit_image_to_window_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.fit_image_to_actual_size_button = customtkinter.CTkButton(self.button_frame, text="Actual Size",
                                                                       command=self.fit_image_to_actual_size)
        self.fit_image_to_actual_size_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.save_file_button = customtkinter.CTkButton(self.button_frame, text="Save as...",
                                                        command=self.save_file)
        self.save_file_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=5)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Image\n"
                                                                            "Next : → right arrow key\n"
                                                                            "Previous : ← left arrow key ",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="Coordinate\n"
                                                                            "Record : left-click\n"
                                                                            "Delete : right-click",
                                                    fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.canvas.bind("<Button-1>", self.on_canvas_left_click)
        self.canvas.bind("<Button-3>", self.on_canvas_right_click)
        self.canvas.bind("<Motion>", self.on_canvas_motion)
        self.canvas.configure(cursor="crosshair")

        self.root.bind("<Right>", self.next_image_keyboard)
        self.root.bind("<Left>", self.previous_image_keyboard)

    def configure_optional_parameter(self):
        fps_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                       " fps (Optional):\n"
                                                       "\n"
                                                       "If you want to add a time column to the output file, "
                                                       "enter a value of fps.",
                                                  title="fps (Optional)")
        self.fps = fps_dialog.get_input()
        if self.fps != "":
            self.fps = float(self.fps)
        self.fps_label.configure(text=f"fps: {self.fps}")
        cm_per_px_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                             " cm/px (Optional):\n"
                                                             "\n"
                                                             "If you want to add physical coordinates (cm) converted "
                                                             "from image coordinates (px) to the output file, "
                                                             "enter a value of cm/px.",
                                                        title="cm/px (Optional)")
        self.cm_per_px = cm_per_px_dialog.get_input()
        if self.cm_per_px != "":
            self.cm_per_px = float(self.cm_per_px)
        self.cm_per_px_label.configure(text=f"cm/px: {self.cm_per_px}")

    def get_dir(self):
        dir_path = customtkinter.filedialog.askdirectory(title='Open Images from Directory ( jpg, png, tif, bmp )')
        path = Path(dir_path)
        image_paths = natsorted([p for p in path.glob('**/*')
                                 if not re.search('Bkg', str(p)) and re.search('/*\.(jpg|jpeg|png|tif|bmp)', str(p))])
        if len(image_paths) != 0:
            interval_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                                "Frame Interval:\n"
                                                                "\n"
                                                                "Examples\n"
                                                                "1: Load all frames (001.jpg, 002.jpg, 003.jpg, ...)\n"
                                                                "2: Load frames every one frame (001,003,005,...)\n"
                                                                "3: Load frames every two frames (001,004,007,...)\n",
                                                           title="Frame Interval")
            self.interval = int(interval_dialog.get_input())
            self.frame_interval_label.configure(text=f"Frame Interval: {self.interval}")
            self.image_paths = image_paths[::self.interval]
            self.configure_optional_parameter()
            self.load_image()

    def get_range(self):
        file_type = [("Supported Files", "*.jpg *.JPG *.jpeg *.png *.PNG *.bmp *.BMP *.tif")]
        image_paths = customtkinter.filedialog.askopenfilenames(title="Open Images by Bounded Selection",
                                                                filetypes=file_type)
        image_paths = natsorted(image_paths)
        if len(image_paths) != 0:
            interval_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                                "Frame Interval:\n"
                                                                "\n"
                                                                "Examples\n"
                                                                "1: Load all frames (001.jpg, 002.jpg, 003.jpg, ...)\n"
                                                                "2: Load frames every one frame (001,003,005,...)\n"
                                                                "3: Load frames every two frames (001,004,007,...)\n",
                                                           title="Frame Interval")
            self.interval = int(interval_dialog.get_input())
            self.frame_interval_label.configure(text=f"Frame Interval: {self.interval}")
            self.image_paths = image_paths[::self.interval]
            self.configure_optional_parameter()
            self.load_image()

    def get_files(self):
        file_type = [("Supported Files", "*.jpg *.JPG *.jpeg *.png *.PNG *.bmp *.BMP *.tif")]
        image_paths = customtkinter.filedialog.askopenfilenames(title="Open Image(s) by Click", filetypes=file_type)
        if len(image_paths) != 0:
            self.image_paths = natsorted(image_paths)
            self.configure_optional_parameter()
            self.load_image()

    def load_image(self, fit_image=False):
        image_path = self.image_paths[self.current_image_index]
        image = Image.open(image_path)
        self.image_file_name = Path(image_path).name
        self.image_width, self.image_height = image.size
        self.image_size_label.configure(text=f"Image Size: {self.image_width} x {self.image_height}")
        self.resized_image = self.resize_image(image, fit_image)
        self.photo_image = ImageTk.PhotoImage(self.resized_image)
        self.canvas.delete("all")
        self.canvas.create_image(0 + self.offset_x, 0 + self.offset_y, anchor=customtkinter.NW, image=self.photo_image)
        self.update_labels(image_path)

    def resize_image(self, image, fit_image=False):
        image_width, image_height = image.size
        win_width = self.image_frame.winfo_width()
        win_height = self.image_frame.winfo_height()
        win_fit_magnification_ratio = min(win_width / image_width, win_height / image_height)
        if self.first_run:
            self.zoom_level = win_fit_magnification_ratio
            self.first_run = False
        if fit_image:
            self.zoom_level = win_fit_magnification_ratio
        new_width = int(image_width * self.zoom_level)
        new_height = int(image_height * self.zoom_level)
        if new_width <= 0 or new_height <= 0:
            return image
        resized_image = image.resize((new_width, new_height), Image.LANCZOS)
        self.resized_image_width = resized_image.width
        self.resized_image_height = resized_image.height
        self.magnification = self.resized_image_width / self.image_width
        return resized_image

    def zoom_image(self):
        self.zoom_level = self.zoom_spinbox.get() * 0.01
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def fit_image_to_window(self):
        fit_image = True
        self.offset_x = 0
        self.offset_y = 0
        self.canvas.delete("all")
        self.load_image(fit_image)
        self.draw_coordinates()

    def fit_image_to_actual_size(self):
        self.zoom_level = 1.0
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def move_image(self, dx, dy):
        self.offset_x += dx
        self.offset_y += dy
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def next_image_keyboard(self, event):
        self.next_image()

    def previous_image_keyboard(self, event):
        self.previous_image()

    def on_canvas_motion(self, event):
        x = (event.x - self.offset_x) / self.magnification
        y = (event.y - self.offset_y) / self.magnification
        self.update_coordinates_label(round(x), round(y))

    def on_canvas_left_click(self, event):
        if self.image_file_name is not None:
            x = (event.x - self.offset_x) / self.magnification
            y = (event.y - self.offset_y) / self.magnification
            self.coordinates.append((x, y))
            self.draw_coordinates()

    def on_canvas_right_click(self, event):
        if self.image_file_name is not None:
            if len(self.coordinates) != 0:
                del self.coordinates[-1]
                self.draw_coordinates()

    def draw_coordinates(self):
        self.canvas.delete("coordinates")
        for coord in self.coordinates:
            x, y = coord
            x = x * self.magnification + self.offset_x
            y = y * self.magnification + self.offset_y
            self.canvas.create_oval(x - 2, y - 2, x + 2, y + 2, fill="red", outline="red", tags="coordinates")

    def record_coordinates(self):
        if len(self.pos) != self.current_image_index:
            self.pos[self.current_image_index] = self.coordinates
            self.file_list[self.current_image_index] = self.image_file_name
        if len(self.pos) == self.current_image_index:
            self.pos.insert(self.current_image_index, self.coordinates)
            self.file_list.insert(self.current_image_index, self.image_file_name)

    def next_image(self):
        if self.current_image_index + 1 == len(self.image_paths):
            self.save_file()
        if self.current_image_index + 1 < len(self.image_paths):
            self.record_coordinates()
            self.current_image_index = (self.current_image_index + 1) % len(self.image_paths)
            self.canvas.delete("all")
            self.load_image()
            if len(self.pos) != self.current_image_index:
                self.coordinates = self.pos[self.current_image_index]
                self.draw_coordinates()
            if len(self.pos) == self.current_image_index:
                self.coordinates = []

    def previous_image(self):
        if self.current_image_index != 0:
            self.current_image_index = (self.current_image_index - 1) % len(self.image_paths)
            self.coordinates = self.pos[self.current_image_index]
            self.canvas.delete("all")
            self.load_image()
            self.draw_coordinates()

    def update_coordinates_label(self, x, y):
        self.coordinates_label.configure(text=f"Coordinates (px): ({x}, {y})")

    def update_labels(self, image_path):
        self.root.title(f"PyCorec {pycorec_version}      [ {self.current_image_index + 1} / {len(self.image_paths)} ]  "
                        f"{Path(image_path).name}")
        self.resized_image_size_label.configure(text=f"Resized Image Size: "
                                                     f"{self.resized_image.width} x {self.resized_image.height}")
        image_magnification = (self.resized_image.width / self.image_width) * 100
        self.image_magnification_label.configure(text=f"Image Magnification (%): {image_magnification:.1f}")
        self.zoom_spinbox.set(image_magnification)

    def save_file(self):
        self.record_coordinates()
        now = datetime.datetime.now()
        current_time = now.strftime("%Y-%m-%d-%H-%M-%S")
        save_path = customtkinter.filedialog.asksaveasfilename(title="Save Coordinates File",
                                                               filetypes=[("Excel Book", ".xlsx"),
                                                                          ("CSV", ".csv")],
                                                               initialfile=f"PyCorec"
                                                                           f"{pycorec_version}_{current_time}",
                                                               defaultextension=".xlsx")
        if ".xlsx" in save_path:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Index"
            ws["B1"] = "File name"
            for i in range(len(self.file_list)):
                ws.cell(row=i + 2, column=1, value=i)
                ws.cell(row=i + 2, column=2, value=self.file_list[i])

            if self.fps != "":
                ws["C1"] = "Time(s)"
                spf = 1 / self.fps
                timestep = self.interval * spf
                end = 0 + (len(self.image_paths) - 1) * timestep
                sec_list = [0 + i * timestep for i in range(int((end - 0) / timestep) + 1)]
                for i in range(len(sec_list)):
                    ws.cell(row=i + 2, column=3, value=sec_list[i])

            flat_pos = [[item for sublist in row for item in sublist] for row in self.pos]
            flat_pos_row_length_check = [len(v) for v in flat_pos]
            flat_pos_row_length = max(flat_pos_row_length_check)
            flat_pos_row_set_length = flat_pos_row_length / 2

            for i in range(1, int(flat_pos_row_set_length + 1)):
                ws.cell(row=1, column=4 + 2 * (i - 1), value=f"x{i}(px)")
                ws.cell(row=1, column=5 + 2 * (i - 1), value=f"y{i}(px)")
            write_list_2d(sheet=ws, l_2d=flat_pos, start_row=2, start_col=4)

            if self.cm_per_px != "":
                cm_pos = [[d * self.cm_per_px for d in row] for row in flat_pos]
                cm_pos = [[-1 * value if column % 2 != 0 else value for column, value in enumerate(row)] for row in cm_pos]
                cm_pos_row_length_check = [len(v) for v in cm_pos]
                cm_pos_row_length = max(cm_pos_row_length_check)
                cm_pos_row_set_length = cm_pos_row_length / 2

                for i in range(1, int(cm_pos_row_set_length + 1)):
                    ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (i - 1), value=f"x{i}(cm)")
                    ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 1 + 2 * (i - 1), value=f"y{i}(cm)")
                write_list_2d(sheet=ws, l_2d=cm_pos, start_row=2, start_col=4 + 2 * (flat_pos_row_set_length + 1 - 1))

                ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1),
                        value=f"xm(px)")
                ws.cell(row=2, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1),
                        value=self.image_width)
                ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 1,
                        value=f"ym(px)")
                ws.cell(row=2, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 1,
                        value=self.image_height)
                ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 2,
                        value=f"xm(cm)")
                ws.cell(row=2, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 2,
                        value=self.image_width * self.cm_per_px)
                ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 3,
                        value=f"ym(cm)")
                ws.cell(row=2, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (cm_pos_row_set_length + 1 - 1) + 3,
                        value=self.image_height * self.cm_per_px * -1)

            font = Font(name="Segoe UI")
            for column in range(1, int(ws.max_column) + 1):
                for row in range(1, int(ws.max_row) + 1):
                    cell = ws.cell(row=row, column=column)
                    cell.font = font

            wb.save(save_path)

        if ".csv" in save_path:
            col_names_list = ["Index", "File name", "Time(s)"]
            index_list = []
            sec_list = []
            for i in range(len(self.file_list)):
                index_list.append(i)

            if self.fps != "":
                spf = 1 / self.fps
                timestep = self.interval * spf
                end = 0 + (len(self.image_paths) - 1) * timestep
                sec_list = [0 + i * timestep for i in range(int((end - 0) / timestep) + 1)]

            col_merge_list = list(zip_longest(index_list, self.file_list, sec_list, fillvalue=None))
            col_merge_list = [list(row) for row in col_merge_list]

            flat_pos = [[item for sublist in row for item in sublist] for row in self.pos]
            flat_pos_row_length_check = [len(v) for v in flat_pos]
            flat_pos_row_length = max(flat_pos_row_length_check)
            flat_pos_row_set_length = flat_pos_row_length / 2
            for i in range(int(flat_pos_row_set_length)):
                col_names_list.append(f"x{i + 1}(px)")
                col_names_list.append(f"y{i + 1}(px)")

            col_merge_list.append(flat_pos)
            pos_merge_list = [item for sublist in col_merge_list for item in sublist]

            if self.cm_per_px != "":
                cm_pos = [[d * self.cm_per_px for d in row] for row in flat_pos]
                cm_pos = [[-1 * value if column % 2 != 0 else value for column, value in enumerate(row)] for row in
                          cm_pos]
                cm_pos_row_length_check = [len(v) for v in cm_pos]
                cm_pos_row_length = max(cm_pos_row_length_check)
                cm_pos_row_set_length = cm_pos_row_length / 2

                for i in range(int(cm_pos_row_set_length)):
                    col_names_list.append(f"x{i + 1}(cm)")
                    col_names_list.append(f"y{i + 1}(cm)")

                pos_merge_list.append(cm_pos)
                pos_merge_list = [item for sublist in pos_merge_list for item in sublist]

                col_names_list.extend([f"xm(px)", f"ym(px)", f"xm(cm)", f"ym(cm)"])
                image_size_list = [self.image_width, self.image_height,
                                   self.image_width * self.cm_per_px, self.image_height * self.cm_per_px * -1]



            with open(save_path, "w", newline="") as file:
                writer = csv.writer(file)
                writer.writerows(save_list)



if __name__ == "__main__":
    app = PyCorec()
    app.root.mainloop()
