import ctypes  # system included
import datetime  # system included
from pathlib import Path  # system included
import re  # system included
import customtkinter  # pip install customtkinter
from natsort import natsorted  # pip install natsort
from openpyxl import Workbook  # pip install openpyxl
from PIL import Image, ImageTk  # pip install pillow


def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])


class PyCorec:
    def __init__(self):
        super().__init__()
        self.magnification = 1
        self.photo_image = None
        self.resized_image = None
        self.image_height = 1
        self.image_width = 1
        self.resized_image_height = 1
        self.resized_image_width = 1
        self.cm_per_px = None
        self.fps = None
        self.interval = 1
        self.image_paths = []
        self.current_image_index = 0
        self.coordinates = []
        self.pos = []
        self.zoom_level = 1.0
        self.offset_x = 0
        self.offset_y = 0

        customtkinter.set_appearance_mode("Dark")

        self.root = customtkinter.CTk()
        self.root.title("PyCorec 2.0.1")

        # configure window size
        user32 = ctypes.windll.user32
        screen_width = user32.GetSystemMetrics(0)
        screen_height = user32.GetSystemMetrics(1) - user32.GetSystemMetrics(3)
        self.root.geometry(f"{screen_width}x{screen_height}+0+0")
        self.root.attributes('-topmost', True)

        self.frame = customtkinter.CTkFrame(self.root)
        self.frame.pack(fill=customtkinter.BOTH, expand=True)

        self.image_frame = customtkinter.CTkFrame(self.frame)
        self.image_frame.pack(side=customtkinter.LEFT, fill=customtkinter.BOTH, expand=True)

        self.canvas = customtkinter.CTkCanvas(self.image_frame, bg="white")
        self.canvas.pack(fill=customtkinter.BOTH, expand=True)

        self.bottom_frame = customtkinter.CTkFrame(self.root)
        self.bottom_frame.pack(side=customtkinter.BOTTOM, fill=customtkinter.X)

        self.label_frame = customtkinter.CTkFrame(self.bottom_frame)
        self.label_frame.pack(side=customtkinter.LEFT, fill=customtkinter.X, padx=10, pady=10)

        # labels
        self.coordinates_label = customtkinter.CTkLabel(self.label_frame, text="Coordinates: (0, 0)")
        self.coordinates_label.pack(side=customtkinter.LEFT)

        self.filename_label = customtkinter.CTkLabel(self.label_frame, text="Filename: ")
        self.filename_label.pack(side=customtkinter.LEFT, padx=10)

        self.image_number_label = customtkinter.CTkLabel(self.label_frame, text="Image Number: ")
        self.image_number_label.pack(side=customtkinter.LEFT, padx=10)

        self.frame_interval_label = customtkinter.CTkLabel(self.label_frame, text="Frame Interval: 1")
        self.frame_interval_label.pack(side=customtkinter.LEFT, padx=10)

        self.image_size_label = customtkinter.CTkLabel(self.label_frame, text="Image Size: ")
        self.image_size_label.pack(side=customtkinter.LEFT, padx=10)

        self.resized_image_size_label = customtkinter.CTkLabel(self.label_frame, text="Resized Image Size: ")
        self.resized_image_size_label.pack(side=customtkinter.LEFT, padx=10)

        self.image_magnification_label = customtkinter.CTkLabel(self.label_frame, text="Image Magnification(%): ")
        self.image_magnification_label.pack(side=customtkinter.LEFT, padx=10)

        self.fps_label = customtkinter.CTkLabel(self.label_frame, text="fps: ")
        self.fps_label.pack(side=customtkinter.LEFT, padx=10)

        self.cm_per_px_label = customtkinter.CTkLabel(self.label_frame, text="cm/px: ")
        self.cm_per_px_label.pack(side=customtkinter.LEFT, padx=10)

        #  buttons
        self.button_frame = customtkinter.CTkFrame(self.frame)
        self.button_frame.pack(side=customtkinter.RIGHT, fill=customtkinter.Y, padx=10, pady=10)

        self.open_image_dir_button = customtkinter.CTkButton(self.button_frame, text="Open Images from Directory",
                                                             command=self.get_dir)
        self.open_image_dir_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_range_button = customtkinter.CTkButton(self.button_frame, text="Open Images by Bounded Selection",
                                                               command=self.get_range)
        self.open_image_range_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.open_image_click_button = customtkinter.CTkButton(self.button_frame, text="Open Image(s) by Click",
                                                               command=self.get_files)
        self.open_image_click_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.next_image_button = customtkinter.CTkButton(self.button_frame, text="Next Image", command=self.next_image)
        self.next_image_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.previous_image_button = customtkinter.CTkButton(self.button_frame, text="Previous Image",
                                                             command=self.previous_image)
        self.previous_image_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.move_up_button = customtkinter.CTkButton(self.button_frame, text="Move Up",
                                                      command=lambda: self.move_image(0, -10))
        self.move_up_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.move_down_button = customtkinter.CTkButton(self.button_frame, text="Move Down",
                                                        command=lambda: self.move_image(0, 10))
        self.move_down_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.move_left_button = customtkinter.CTkButton(self.button_frame, text="Move Left",
                                                        command=lambda: self.move_image(-10, 0))
        self.move_left_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.move_right_button = customtkinter.CTkButton(self.button_frame, text="Move Right",
                                                         command=lambda: self.move_image(10, 0))
        self.move_right_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.fit_image_to_window_button = customtkinter.CTkButton(self.button_frame, text="Fit Image to Window",
                                                                  command=self.fit_image_to_window)
        self.fit_image_to_window_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.fit_image_to_actual_size_button = customtkinter.CTkButton(self.button_frame,
                                                                       text="Fit Image to Actual Size",
                                                                       command=self.fit_image_to_actual_size)
        self.fit_image_to_actual_size_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.blank_button = customtkinter.CTkButton(self.button_frame, text="", fg_color="transparent", hover=False)
        self.blank_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.save_xlsx_button = customtkinter.CTkButton(self.button_frame,
                                                        text="Save  as Excel File",
                                                        command=self.save_xlsx)
        self.save_xlsx_button.pack(fill=customtkinter.X, padx=10, pady=10)

        self.canvas.bind("<Button-1>", self.on_canvas_left_click)
        self.canvas.bind("<Button-3>", self.on_canvas_right_click)
        self.canvas.bind("<MouseWheel>", self.on_mouse_wheel)
        self.canvas.bind("<Motion>", self.on_canvas_motion)
        self.canvas.configure(cursor="crosshair")

        self.root.bind("<Right>", self.next_image_keyboard)
        self.root.bind("<Left>", self.previous_image_keyboard)

    def configure_optional_parameter(self):
        fps_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                       " fps (Optional):\n"
                                                       "\n"
                                                       "If you want to add a time column to the output file, "
                                                       "enter a value of fps.",
                                                  title="fps (Optional)")
        self.fps = fps_dialog.get_input()
        if self.fps != "":
            self.fps = float(self.fps)
        self.fps_label.configure(text=f"fps: {self.fps}")
        cm_per_px_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                             " cm/px (Optional):\n"
                                                             "\n"
                                                             "If you want to add physical coordinates (cm) converted "
                                                             "from image coordinates (px) to the output file, "
                                                             "enter a value of cm/px.",
                                                        title="cm/px (Optional)")
        self.cm_per_px = cm_per_px_dialog.get_input()
        if self.cm_per_px != "":
            self.cm_per_px = float(self.cm_per_px)
        self.cm_per_px_label.configure(text=f"cm/px: {self.cm_per_px}")

    def get_dir(self):
        dir_path = customtkinter.filedialog.askdirectory(title='Open Images from Directory')
        path = Path(dir_path)
        image_paths = natsorted([p for p in path.glob('**/*')
                                 if not re.search('Bkg', str(p)) and re.search('/*\.(jpg|jpeg|png|tif|bmp)', str(p))])
        interval_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                            "Frame Interval:\n"
                                                            "\n"
                                                            "Examples\n"
                                                            "1: Load all frames (001.jpg, 002.jpg, 003.jpg, ...)\n"
                                                            "2: Load frames every one frame (001,003,005,...)\n"
                                                            "3: Load frames every two frames (001,004,007,...)\n",
                                                       title="Frame Interval")
        self.interval = int(interval_dialog.get_input())
        self.frame_interval_label.configure(text=f"Frame Interval: {self.interval}")
        self.image_paths = image_paths[::self.interval]
        self.configure_optional_parameter()
        self.load_image()

    def get_range(self):
        file_type = [("Supported Files", "*.jpg *.JPG *.jpeg *.png *.PNG *.bmp *.BMP *.tif")]
        image_paths = customtkinter.filedialog.askopenfilenames(title="Open Images by Bounded Selection",
                                                                filetypes=file_type)
        image_paths = natsorted(image_paths)
        interval_dialog = customtkinter.CTkInputDialog(text="Input\n"
                                                            "Frame Interval:\n"
                                                            "\n"
                                                            "Examples\n"
                                                            "1: Load all frames (001.jpg, 002.jpg, 003.jpg, ...)\n"
                                                            "2: Load frames every one frame (001,003,005,...)\n"
                                                            "3: Load frames every two frames (001,004,007,...)\n",
                                                       title="Frame Interval")
        self.interval = int(interval_dialog.get_input())
        self.frame_interval_label.configure(text=f"Frame Interval: {self.interval}")
        self.image_paths = image_paths[::self.interval]
        self.configure_optional_parameter()
        self.load_image()

    def get_files(self):
        file_type = [("Supported Files", "*.jpg *.JPG *.jpeg *.png *.PNG *.bmp *.BMP *.tif")]
        image_paths = customtkinter.filedialog.askopenfilenames(title="Open Image(s) by Click", filetypes=file_type)
        self.image_paths = natsorted(image_paths)
        self.configure_optional_parameter()
        self.load_image()

    def load_image(self, actual_size=False):
        image_path = self.image_paths[self.current_image_index]
        image = Image.open(image_path)
        self.image_width, self.image_height = image.size
        self.image_size_label.configure(text=f"Image Size: {self.image_width} x {self.image_height}")
        self.resized_image = self.resize_image(image, actual_size)
        self.photo_image = ImageTk.PhotoImage(self.resized_image)
        self.canvas.delete("all")
        self.canvas.create_image(0 + self.offset_x, 0 + self.offset_y, anchor=customtkinter.NW, image=self.photo_image)
        self.update_labels(image_path)

    def resize_image(self, image, actual_size):
        width, height = image.size
        max_width = self.image_frame.winfo_width()
        max_height = self.image_frame.winfo_height()
        if actual_size:
            new_width = width
            new_height = height
        else:
            magnification_ratio = min(max_width / width, max_height / height)
            new_width = int(width * magnification_ratio * self.zoom_level)
            new_height = int(height * magnification_ratio * self.zoom_level)

            if new_width <= 0 or new_height <= 0:
                return image

        resized_image = image.resize((new_width, new_height), Image.LANCZOS)
        self.resized_image_width = resized_image.width
        self.resized_image_height = resized_image.height
        self.magnification = self.resized_image_width / self.image_width
        return resized_image

    def zoom_image(self, zoom_in):
        if zoom_in:
            self.zoom_level *= 1.1
        else:
            self.zoom_level /= 1.1

        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def fit_image_to_window(self):
        self.zoom_level = 1.0
        self.offset_x = 0
        self.offset_y = 0
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def fit_image_to_actual_size(self):
        actual_size = True
        self.offset_x = 0
        self.offset_y = 0
        self.canvas.delete("all")
        self.load_image(actual_size)
        self.draw_coordinates()

    def move_image(self, dx, dy):
        self.offset_x += dx
        self.offset_y += dy
        self.canvas.delete("all")
        self.load_image()
        self.draw_coordinates()

    def next_image_keyboard(self, event):
        self.next_image()

    def previous_image_keyboard(self, event):
        self.previous_image()

    def on_mouse_wheel(self, event):
        if event.delta > 0:
            self.zoom_image(True)
        else:
            self.zoom_image(False)

    def on_canvas_motion(self, event):
        x = (event.x - self.offset_x) / self.magnification
        y = (event.y - self.offset_y) / self.magnification
        self.update_coordinates_label(round(x), round(y))

    def on_canvas_left_click(self, event):
        x = (event.x - self.offset_x) / self.magnification
        y = (event.y - self.offset_y) / self.magnification
        self.coordinates.append((x, y))
        print(self.coordinates)
        self.draw_coordinates()

    def on_canvas_right_click(self, event):
        if len(self.coordinates) != 0:
            del self.coordinates[-1]
        print(self.coordinates)
        self.draw_coordinates()

    def draw_coordinates(self):
        self.canvas.delete("coordinates")
        for coord in self.coordinates:
            x, y = coord
            x = x * self.magnification + self.offset_x
            y = y * self.magnification + self.offset_y
            self.canvas.create_oval(x - 2, y - 2, x + 2, y + 2, fill="red", outline="red", tags="coordinates")

    def record_coordinates(self):
        self.pos.insert(self.current_image_index, self.coordinates)

    def next_image(self):
        if self.current_image_index + 1 == len(self.image_paths):
            self.save_xlsx()
        if self.current_image_index + 1 < len(self.image_paths):
            self.record_coordinates()
            self.coordinates = []
            self.current_image_index = (self.current_image_index + 1) % len(self.image_paths)
            self.canvas.delete("all")
            self.load_image()

    def previous_image(self):
        self.current_image_index = (self.current_image_index - 1) % len(self.image_paths)
        self.coordinates = self.pos[self.current_image_index]
        self.canvas.delete("all")
        self.load_image()

    def update_coordinates_label(self, x, y):
        self.coordinates_label.configure(text=f"Coordinates: ({x}, {y})")

    def update_labels(self, image_path):
        self.filename_label.configure(text=f"Filename: {Path(image_path).name}")
        self.image_number_label.configure(text=f"Image Number: {self.current_image_index + 1}/{len(self.image_paths)}")
        self.resized_image_size_label.configure(text=f"Resized Image Size: "
                                                     f"{self.resized_image.width} x {self.resized_image.height}")
        image_magnification = (self.resized_image.width / self.image_width) * 100
        self.image_magnification_label.configure(text=f"Image Magnification(%): {image_magnification:.1f}")

    def save_xlsx(self):
        self.record_coordinates()
        now = datetime.datetime.now()
        current_time = now.strftime("%Y-%m-%d-%H-%M-%S")
        xlsx_path = customtkinter.filedialog.asksaveasfilename(title="Save Coordinates File",
                                                               filetypes=[("Excel Book", ".xlsx")],
                                                               initialfile=f"PyCorec2.0.1_{current_time}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Index"
        for i in range(len(self.image_paths)):
            ws.cell(row=i + 2, column=1, value=i)
        if self.fps != "":
            ws["B1"] = "Time(s)"
            spf = 1 / self.fps
            timestep = self.interval * spf
            end = 0 + (len(self.image_paths) - 1) * timestep
            sec = [0 + i * timestep for i in range(int((end - 0) / timestep) + 1)]
            for i in range(len(sec)):
                ws.cell(row=i + 2, column=2, value=sec[i])
        flat_pos = [[item for sublist in row for item in sublist] for row in self.pos]
        flat_pos_row_length_check = [len(v) for v in flat_pos]
        flat_pos_row_length = max(flat_pos_row_length_check)
        flat_pos_row_set_length = flat_pos_row_length / 2
        for i in range(1, int(flat_pos_row_set_length + 1)):
            ws.cell(row=1, column=3 + 2 * (i - 1), value=f"x{i}(px)")
            ws.cell(row=1, column=4 + 2 * (i - 1), value=f"y{i}(px)")
        write_list_2d(sheet=ws, l_2d=flat_pos, start_row=2, start_col=3)
        if self.cm_per_px != "":
            cm_pos = [[d * self.cm_per_px for d in row] for row in flat_pos]
            cm_pos = [[-1 * value if column % 2 != 0 else value for column, value in enumerate(row)] for row in cm_pos]
            cm_pos_row_length_check = [len(v) for v in cm_pos]
            cm_pos_row_length = max(cm_pos_row_length_check)
            cm_pos_row_set_length = cm_pos_row_length / 2
            for i in range(1, int(cm_pos_row_set_length + 1)):
                ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 2 * (i - 1), value=f"x{i}(cm)")
                ws.cell(row=1, column=(4 + 2 * (flat_pos_row_set_length + 1 - 1)) + 1 + 2 * (i - 1), value=f"y{i}(cm)")
            write_list_2d(sheet=ws, l_2d=cm_pos, start_row=2, start_col=4 + 2 * (flat_pos_row_set_length + 1 - 1))
        wb.save(xlsx_path)


if __name__ == "__main__":
    app = PyCorec()
    app.root.mainloop()
